import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from dotenv import load_dotenv
import os
import time

planilha_controle = r"caminho_da_planilha\exemplo.xlsx" 

# Carrega variáveis do .env
load_dotenv()
EMAIL = os.getenv("EMAIL")
SENHA = os.getenv("SENHA")

def limpar_downloads(caminho_pasta):
    try:
        arquivos = os.listdir(caminho_pasta)
        for arquivo in arquivos:
            caminho_arquivo = os.path.join(caminho_pasta, arquivo)
            if os.path.isfile(caminho_arquivo):
                os.remove(caminho_arquivo)
    except Exception as e:
        print(f"Erro ao limpar a pasta de downloads: {e}")

def iniciar_navegador():
    options = Options()
    options.add_argument("--start-maximized")

    # Suprime logs do DevTools e erros de depuração
    options.add_argument("--log-level=3")  # 0=INFO, 1=WARNING, 2=LOG_ERROR, 3=FATAL
    options.add_experimental_option("excludeSwitches", ["enable-logging"])

    options.add_experimental_option("prefs", {
        "download.prompt_for_download": False,
        "download.default_directory": os.path.join(os.getcwd(), "downloads"),
        "safebrowsing.enabled": True
    })

    service = Service("./msedgedriver.exe")
    driver = webdriver.Edge(service=service, options=options)
    return driver

def realizar_login(driver, wait):
    driver.get("https://dmi.isso.digital/?p=Login&e=Cg==&u=Lw==")
    try:
        wait.until(EC.presence_of_element_located((By.ID, "username"))).send_keys(EMAIL)
        driver.find_element(By.ID, "password").send_keys(SENHA)
        driver.find_element(By.XPATH, "//input[@value='ENTRAR']").click()
        print("Login enviado.")
        return True
    except TimeoutException:
        print("Erro: campos de login não encontrados.")
        return False

def acessar_datalog(driver, wait):
    try:
        datalog_botao = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[div[text()='Datalog']]")))
        datalog_botao.click()
        print("Datalog acessado.")
    except TimeoutException:
        print("Erro: botão Datalog não encontrado.")

def acessar_graficos(driver, wait):
    try:
        graficos_aba = wait.until(EC.element_to_be_clickable((By.ID, "lnk-Aba3")))
        graficos_aba.click()
        print("Aba Gráficos acessada com sucesso.")
    except TimeoutException:
        print("Erro: aba 'Gráficos' não acessada.")

def fechar_popups(driver, wait):
    for i in range(2):
        try:
            # Aguarda o botão aparecer no DOM
            botao = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "modal-fecha-alerta")))

            # Aguarda até que o botão esteja visível e desbloqueado
            wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "modal-fecha-alerta")))

            # Força o clique via JavaScript para evitar bloqueios visuais
            driver.execute_script("arguments[0].click();", botao)
            time.sleep(1)  # pequena pausa entre os cliques
        except TimeoutException:
            print(f"Popup {2 - i} não apareceu ou não pôde ser fechado.")
            break

def exportar_graficos(driver, wait, id_base="525517"):
    try:
        # Scroll até o gráfico principal
        grafico_principal = wait.until(
            EC.presence_of_element_located((By.XPATH, f'//*[@id="box_{id_base}"]'))
        )
        ActionChains(driver).move_to_element(grafico_principal).perform()
        # Selecionar "Período personalizado"
        wait.until(EC.element_to_be_clickable((By.XPATH, f'//*[@id="aux_{id_base}"]/div[2]/select/option[9]'))).click()
        time.sleep(1)
        # Clicar no botão "Opções"
        wait.until(EC.element_to_be_clickable((By.XPATH, f'//*[@id="aux_{id_base}"]/div[2]/button[1]'))).click()
        time.sleep(1)
        # Selecionar agrupamento "Horária"
        wait.until(EC.element_to_be_clickable((By.XPATH, f'//*[@id="aux_{id_base}"]/div[2]/div/div[2]/select/option[2]'))).click()
        # Clicar no botão "Carregar" para atualizar o gráfico
        wait.until(EC.element_to_be_clickable((By.XPATH, f'//*[@id="aux_{id_base}"]/div[2]/button[2]'))).click()
        time.sleep(5)  # aguarda renderização do gráfico
        # Clicar no botão "Sincronizar" para aplicar os filtros aos demais gráficos
        wait.until(EC.element_to_be_clickable((By.XPATH, f'//*[@id="aux_{id_base}"]/div[1]/div[2]'))).click()
        time.sleep(7)  # aguarda sincronização

        # Exportar o gráfico principal (fase-neutro)
        wait.until(EC.element_to_be_clickable((By.XPATH, f'//*[@id="aux_{id_base}"]/div[1]/button[1]'))).click()
        time.sleep(1)
        wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="div_form"]/div[1]/button[1]'))).click()
        time.sleep(5)  # aguarda o download
        wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[7]/div/div[3]'))).click()
        time.sleep(1)

        # Exportar o segundo gráfico (fase-fase)
        id_segundo = "525518"
        # Scroll até o gráfico
        grafico_segundo = wait.until(
            EC.presence_of_element_located((By.XPATH, f'//*[@id="box_{id_segundo}"]'))
        )
        ActionChains(driver).move_to_element(grafico_segundo).perform()
        # Clicar no botão de exportar
        wait.until(EC.element_to_be_clickable((By.XPATH, f'//*[@id="aux_{id_segundo}"]/div[1]/button[1]'))).click()
        time.sleep(1)
        # Confirmar exportação no popup
        wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="div_form"]/div[1]/button[1]'))).click()
        time.sleep(5)
        wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[7]/div/div[3]'))).click()
        time.sleep(1)

        # Exportar o terceiro gráfico (correntes)
        id_terceiro = "525516"
        # Scroll até o gráfico
        grafico_terceiro = wait.until(
            EC.presence_of_element_located((By.XPATH, f'//*[@id="box_{id_terceiro}"]'))
        )
        ActionChains(driver).move_to_element(grafico_terceiro).perform()
        # Clicar no botão de exportar
        wait.until(EC.element_to_be_clickable((By.XPATH, f'//*[@id="aux_{id_terceiro}"]/div[1]/button[1]'))).click()
        time.sleep(1)
        # Confirmar exportação no popup
        wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="div_form"]/div[1]/button[1]'))).click()
        time.sleep(5)
        # Fechar o popup
        wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[7]/div/div[3]'))).click()
        time.sleep(1)

    except Exception as e:
        print(f"Erro ao aplicar agrupamento: {e}")

def fase_neutro(caminho_arquivo):
    try:
        # Carrega o CSV com separador e encoding apropriado
        df = pd.read_csv(caminho_arquivo, sep=",", encoding="utf-8")

        # Converte coluna de data para datetime
        df["Data"] = pd.to_datetime(df["Data"], format="%d/%m/%Y %H:%M:%S", errors="coerce")

        # Converte tensões para float (substitui vírgula por ponto)
        for coluna in ["Tensão A", "Tensão B", "Tensão C"]:
            df[coluna] = df[coluna].astype(str).str.replace(",", ".").astype(float)

        # Filtra apenas os dados do dia anterior
        ontem = (datetime.now() - timedelta(days=1)).date()
        df_filtrado = df[df["Data"].dt.date == ontem]

        # Remove a coluna de data e o cabeçalho
        dados_sem_data = df_filtrado[["Tensão A", "Tensão B", "Tensão C"]].values.tolist()
        return dados_sem_data

    except Exception as e:
        print(f"Erro ao preparar dados: {e}")
        return []
    
def fase_fase(caminho_arquivo):
    try:
        df = pd.read_csv(caminho_arquivo, sep=",", encoding="utf-8")

        df["Data"] = pd.to_datetime(df["Data"], format="%d/%m/%Y %H:%M:%S", errors="coerce")

        for coluna in ["Tensão A-B", "Tensão B-C", "Tensão C-A"]:
            df[coluna] = df[coluna].astype(str).str.replace(",", ".").astype(float)

        ontem = (datetime.now() - timedelta(days=1)).date()
        df_filtrado = df[df["Data"].dt.date == ontem]
        dados_sem_data = df_filtrado[["Tensão A-B", "Tensão B-C", "Tensão C-A"]].values.tolist()
        return dados_sem_data

    except Exception as e:
        print(f"Erro ao preparar dados fase-fase: {e}")
        return []

def correntes(caminho_arquivo):
    try:
        df = pd.read_csv(caminho_arquivo, sep=",", encoding="utf-8")

        df["Data"] = pd.to_datetime(df["Data"], format="%d/%m/%Y %H:%M:%S", errors="coerce")

        for coluna in ["Corrente A", "Corrente B", "Corrente C"]:
            df[coluna] = (
                df[coluna]
                .astype(str)
                .str.replace(".", "", regex=False)  # remove separador de milhar
                .str.replace(",", ".", regex=False)  # converte decimal
                .astype(float)
            )
        ontem = (datetime.now() - timedelta(days=1)).date()
        df_filtrado = df[df["Data"].dt.date == ontem]

        dados_sem_data = df_filtrado[["Corrente A", "Corrente B", "Corrente C"]].values.tolist()
        return dados_sem_data

    except Exception as e:
        print(f"Erro ao preparar dados de corrente: {e}")
        return []

def colar_dados(caminho_planilha, tensao_fn, tensao_ff, corrente):
    try:
        wb = load_workbook(caminho_planilha)
        ws = wb.active

        # Determina a primeira linha livre com base na coluna I
        linha = ws.max_row + 1
        while ws[f"I{linha}"].value is None and linha > 1:
            linha -= 1
        linha += 1

        # Garante que todas as listas tenham o mesmo número de linhas
        total_linhas = min(len(tensao_fn), len(tensao_ff), len(corrente))

        #Verificar as colunas desejadas para gravação dos dados!!
        for i in range(total_linhas):
            # Fase-neutro → I, J, K
            ws[f"I{linha+i}"] = tensao_fn[i][0]
            ws[f"J{linha+i}"] = tensao_fn[i][1]
            ws[f"K{linha+i}"] = tensao_fn[i][2]

            # Fase-fase → L, M, N
            ws[f"L{linha+i}"] = tensao_ff[i][0]
            ws[f"M{linha+i}"] = tensao_ff[i][1]
            ws[f"N{linha+i}"] = tensao_ff[i][2]

            # Correntes → O, P, Q
            ws[f"O{linha+i}"] = corrente[i][0]
            ws[f"P{linha+i}"] = corrente[i][1]
            ws[f"Q{linha+i}"] = corrente[i][2]

        wb.save(caminho_planilha)
    except Exception as e:
        print(f"Erro ao colar dados na planilha: {e}")

def main():
    limpar_downloads("./downloads")
    driver = iniciar_navegador()
    wait = WebDriverWait(driver, 20)
    if not realizar_login(driver, wait):
        driver.quit()
        return
    # Aguarda e fecha aba extra pós-login
    wait.until(lambda d: len(d.window_handles) > 1)
    abas = driver.window_handles
    driver.switch_to.window(abas[1])
    driver.close()
    driver.switch_to.window(abas[0])

    acessar_datalog(driver, wait)

    # Aguarda e troca para aba do Datalog
    wait.until(lambda d: len(d.window_handles) > 1)
    abas = driver.window_handles
    driver.switch_to.window(abas[0])
    driver.close()
    driver.switch_to.window(abas[1])

    # Acessa aba Gráficos
    acessar_graficos(driver, wait)
    time.sleep(30)
    fechar_popups(driver,wait)
    exportar_graficos(driver, wait)
    time.sleep(3)
    driver.quit()

    dados_fase_neutro = fase_neutro("./downloads/tensões-fase---neutro.csv")
    dados_fase_fase = fase_fase("./downloads/tensões-fase---fase.csv")
    dados_corrente = correntes("./downloads/correntes.csv")
    colar_dados(planilha_controle, dados_fase_neutro, dados_fase_fase, dados_corrente)

if __name__ == "__main__":
    main()